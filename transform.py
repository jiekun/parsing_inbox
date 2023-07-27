from openpyxl import load_workbook


def main():
    inbox = load_workbook(filename='inbox.xlsx')
    inbox_data = inbox['Sheet1']

    parsed_inbox_data = []

    # parse data row by row. start from 6th row
    for row in inbox_data.iter_rows(min_row=6, max_col=28):
        # validation check
        if (
                row[0].value is None or
                row[2].value is None
        ):
            continue

        # build a data map that represent the current row
        parsed_inbox_data.append({
            'index': str(row[0].value) if row[0].value else '',
            'issue_clue_index': str(row[1].value) if row[1].value else '',
            'visitor_name': str(row[2].value) if row[2].value else '',
            'visit_date': str(row[3].value) if row[3].value else '',
            'visitor_job': str(row[4].value) if row[4].value else '',
            'visitor_id': str(row[5].value) if row[5].value else '',
            'visitor_phone': str(row[6].value) if row[6].value else '',
            'receptionist_name': str(row[7].value) if row[7].value else '',
            'visit_method': str(row[8].value) if row[8].value else '',
            'visit_type': str(row[9].value) if row[9].value else '',
            'visitor_num': str(row[10].value) if row[10].value else '',
            'duplicate_num': str(row[11].value) if row[11].value else '',
            'reported_name': str(row[12].value) if row[12].value else '',
            'reported_job': str(row[13].value) if row[13].value else '',
        })

        print(parsed_inbox_data)

    # generate a new one based on template
    template = load_workbook(filename='template.xlsx')
    for each_inbox_data in parsed_inbox_data:
        template['Sheet1']['B3'].value = each_inbox_data.get('visit_date')
        template['Sheet1']['E3'].value = each_inbox_data.get('visit_method')
        template['Sheet1']['C4'].value = each_inbox_data.get('visitor_name')
        template['Sheet1']['E4'].value = each_inbox_data.get('visitor_phone')
        template['Sheet1']['C5'].value = each_inbox_data.get('visitor_id')
        template['Sheet1']['C6'].value = each_inbox_data.get('reported_name')
        template['Sheet1']['E7'].value = each_inbox_data.get('reported_job')
        template.save(each_inbox_data.get('index') + '.xlsx')


if __name__ == '__main__':
    main()
